import datetime
import io
from fastapi import APIRouter, HTTPException, Query, Depends, Response
from sqlalchemy.orm import Session
from sqlalchemy import func

from src.database.session import SessionLocal
from src.database.models import Asset, ChargeDischargePlan, BessTelemetry, PriceForecast
from src.modules.reporting_service.services import ReportingService
from src.modules.reporting_service.forecast_accuracy import compute_rolling_accuracy, get_profit_capture_ratio
from src.core.security import RoleChecker
from src.api.v1.endpoints.optimization import get_manual_overrides
from src.api.v1.endpoints.forecast import get_actual_prices

router = APIRouter()

@router.get("/executive-summary", dependencies=[Depends(RoleChecker(["Viewer", "Operator", "Manager", "Admin"]))])
async def get_executive_summary(
    asset_id: str = Query(..., description="UUID of the BESS asset"),
    period: str = Query("month", description="One of day, week, month, year")
):
    db = SessionLocal()
    try:
        report = ReportingService.get_executive_summary_report(db, asset_id)
        return report
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error compiling executive summary: {str(e)}")
    finally:
        db.close()

@router.get("/forecast-accuracy", dependencies=[Depends(RoleChecker(["Viewer", "Operator", "Manager", "Admin"]))])
async def get_forecast_accuracy(
    days: int = Query(30, description="Скільки останніх днів порівняти прогноз/факт")
):
    db = SessionLocal()
    try:
        live = compute_rolling_accuracy(db, days=days)
        ratio = get_profit_capture_ratio(db)
        return {"live_accuracy": live, "profit_capture_ratio": ratio}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error computing forecast accuracy: {str(e)}")
    finally:
        db.close()

@router.get("/market-conditions", dependencies=[Depends(RoleChecker(["Viewer", "Operator", "Manager", "Admin"]))])
async def get_market_conditions():
    """
    Реальний операційний знімок "на зараз" для панелі диспетчера: остання
    зібрана ціна газу, транскордонний нетто-експорт (ENTSO-E) та keyword-сигнал
    з публічних каналів Укренерго/Міненерго. Замінює мертві повзунки
    "Ринкові фактори прогнозування", які раніше нічого насправді не міняли.
    """
    import os
    import pandas as pd
    from src.core.config import settings
    import src.modules.external_data_service.telegram_public as ext_tg

    result = {
        "gas_price_eur_mwh": None,
        "gas_price_as_of": None,
        "grid_net_export_mw": None,
        "grid_net_export_as_of": None,
        "grid_stress_today": {"grid_stress_high": 0, "grid_stress_medium": 0, "mentions": 0},
        "latest_posts": [],
    }

    csv_path = os.path.join(settings.DATA_DIR, "historical_data_merged.csv")
    if os.path.exists(csv_path):
        try:
            df = pd.read_csv(csv_path, usecols=["Datetime", "Gas_Price_EUR_MWh", "Grid_Net_Export_MW"])
            df["Datetime"] = pd.to_datetime(df["Datetime"])

            gas_series = df.dropna(subset=["Gas_Price_EUR_MWh"])
            if not gas_series.empty:
                last = gas_series.iloc[-1]
                result["gas_price_eur_mwh"] = float(last["Gas_Price_EUR_MWh"])
                result["gas_price_as_of"] = last["Datetime"].isoformat()

            flow_series = df.dropna(subset=["Grid_Net_Export_MW"])
            if not flow_series.empty:
                last = flow_series.iloc[-1]
                result["grid_net_export_mw"] = float(last["Grid_Net_Export_MW"])
                result["grid_net_export_as_of"] = last["Datetime"].isoformat()
        except Exception as e:
            result["error"] = f"Error reading market conditions from CSV: {str(e)}"

    today_str = datetime.datetime.utcnow().date().isoformat()
    stress = ext_tg.daily_grid_stress_signal()
    if today_str in stress:
        result["grid_stress_today"] = stress[today_str]

    result["latest_posts"] = ext_tg.get_latest_posts(n=3)

    return result

@router.get("/export-day", dependencies=[Depends(RoleChecker(["Viewer", "Operator", "Manager", "Admin"]))])
async def export_day_excel(asset_id: str, date: str):
    """
    Погодинний Excel-звіт за добу: прогноз ціни, факт ціни (якщо є),
    заряд/розряд і ціна виконання (реальний заявочний профіль — та сама
    логіка, що й get_manual_overrides, щоб цифри в файлі завжди збігалися з
    тим, що диспетчер бачить на екрані). Факт ціни — та сама логіка, що й
    get_actual_prices (спершу локальна БД, інакше живий запит до
    oree.com.ua, якщо доба вже опублікована, але кеш ще не досинхронізовано —
    без цього факт міг бути порожнім навіть коли ціна вже реально відома).
    Година підписана 1-24 (не 0-23), як на самому oree.com.ua.

    Різниця Факт-Прогноз показує напрям і величину помилки прогнозу за цю
    конкретну добу (позитивна — факт вищий за прогноз). Заряд/Розряд —
    крайні праві стовпці, з міні-графіком просто в самій колонці (Excel
    Data Bars) — не окрема діаграма збоку, а "стовпчик заряду батареї"
    прямо в комірках, як просив диспетчер.

    Саме .xlsx, а не .csv — у CSV Excel сам вгадує типи комірок при
    відкритті (напр. "1-2" перетворюється на дату) без жодного способу це
    заборонити; у .xlsx тип кожної комірки заданий явно, Excel нічого не
    вгадує.
    """
    try:
        datetime.datetime.strptime(date, '%Y-%m-%d')
    except ValueError:
        raise HTTPException(status_code=400, detail="date має бути у форматі YYYY-MM-DD")

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import DataBarRule

    db = SessionLocal()
    try:
        asset = db.query(Asset).filter(Asset.id == asset_id).first()
        if not asset:
            raise HTTPException(status_code=404, detail="Asset not found")

        target_dt = datetime.datetime.strptime(date, '%Y-%m-%d')

        forecasts = db.query(PriceForecast).filter(
            PriceForecast.forecast_run_at == target_dt
        ).order_by(PriceForecast.timestamp).all()
        forecast_by_hour = {f.timestamp.hour: f.predicted_price_uah for f in forecasts}

        dispatch = await get_manual_overrides(asset_id=asset_id, date=date)
        dispatch_by_hour = {o["hour"]: o for o in dispatch["overrides"]}
        power_limit_mw = asset.power_mw
    finally:
        db.close()

    actual = await get_actual_prices(target_date=date)
    actual_by_hour = (
        dict(zip(actual["hours"], actual["actual_prices_uah"])) if actual.get("available") else {}
    )

    wb = Workbook()
    ws = wb.active
    ws.title = date

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")
    title_font = Font(bold=True, size=13)

    ws["A1"] = f"SmartBESS EMS — погодинний звіт за {date}"
    ws["A1"].font = title_font
    ws["A2"] = f"Актив: {asset.name}"
    ws["A2"].font = Font(italic=True, color="6B7280")

    headers = [
        "Година", "Прогноз ціни, ₴/МВт·год", "Факт ціни, ₴/МВт·год", "Різниця Факт-Прогноз, ₴/МВт·год",
        "Ціна виконання, ₴/МВт·год", "Сума, ₴", "Ручна корекція", "Заряд, МВт", "Розряд, МВт",
    ]
    header_row = 4
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for hour in range(24):
        row = header_row + 1 + hour
        d = dispatch_by_hour.get(hour, {})
        power_mw = d.get("power_mw", 0.0) or 0.0
        charge_mw = -power_mw if power_mw < 0 else 0.0
        discharge_mw = power_mw if power_mw > 0 else 0.0
        was_active = power_mw != 0.0

        # Година підписана 1-24 (oree.com.ua "Година 1" = 00:00-01:00), не 0-23.
        ws.cell(row=row, column=1, value=hour + 1).number_format = "0"
        fc = forecast_by_hour.get(hour)
        ws.cell(row=row, column=2, value=round(fc, 2) if fc is not None else None).number_format = "#,##0.00"
        ac = actual_by_hour.get(hour)
        ws.cell(row=row, column=3, value=round(ac, 2) if ac is not None else None).number_format = "#,##0.00"
        diff = (ac - fc) if (ac is not None and fc is not None) else None
        ws.cell(row=row, column=4, value=round(diff, 2) if diff is not None else None).number_format = "+#,##0.00;-#,##0.00"

        # Ціна виконання/Сума мають сенс лише в годинах, де реально відбувся
        # заряд чи розряд — без активності це просто дублювало б прогноз
        # (диспетчер: "мало інформативно"). Сума = ціна × обсяг (₴/МВт·год ×
        # МВт × 1г) зі знаком: від'ємна — витрати на заряд, додатна — дохід
        # від розряду.
        price_uah = d.get("price_uah")
        if was_active and price_uah is not None:
            ws.cell(row=row, column=5, value=round(price_uah, 2)).number_format = "#,##0.00"
            suma = price_uah * power_mw
            ws.cell(row=row, column=6, value=round(suma, 2)).number_format = "+#,##0.00;-#,##0.00"
        else:
            ws.cell(row=row, column=5, value=None)
            ws.cell(row=row, column=6, value=None)

        ws.cell(row=row, column=7, value="так" if d.get("is_overridden") else "ні").alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=8, value=round(charge_mw, 3)).number_format = "#,##0.000"
        ws.cell(row=row, column=9, value=round(discharge_mw, 3)).number_format = "#,##0.000"

    widths = [10, 20, 18, 22, 22, 14, 14, 12, 12]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    last_row = header_row + 24
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # Міні-графік заряду/розряду прямо в комірках крайніх правих колонок
    # (Excel Data Bars), а не окрема діаграма збоку — "стовпчик заряду
    # батареї" в самому стовпці, як і просив диспетчер.
    charge_range = f"H{header_row + 1}:H{last_row}"
    discharge_range = f"I{header_row + 1}:I{last_row}"
    charge_rule = DataBarRule(start_type="num", start_value=0, end_type="num", end_value=power_limit_mw, color="3B82F6", showValue=True)
    discharge_rule = DataBarRule(start_type="num", start_value=0, end_type="num", end_value=power_limit_mw, color="059669", showValue=True)
    ws.conditional_formatting.add(charge_range, charge_rule)
    ws.conditional_formatting.add(discharge_range, discharge_rule)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"smartbess_{date}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
